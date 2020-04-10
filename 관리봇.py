import discord, asyncio
import datetime
import openpyxl
import os

client = discord.Client()


@client.event
async def on_ready():
    print("login")
    print(client.user.name)
    print(client.user.id)
    await client.change_presence(status=discord.Status.online, activity=discord.Game("관리 "))

@client.event
async def on_message(message):
    if message.content.startswith("!정보"):
        date = datetime.datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000) / 1000)
        embed = discord.Embed(color=0x00ff00)
        embed.add_field(name="이름", value=message.author.name, inline=True)
        embed.add_field(name="서버닉넴", value=message.author.display_name, inline=True)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일", inline=True)
        embed.add_field(name="아이디", value=message.author.id, inline=True)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(embed=embed)
    
    if message.content == "!투표":
        vote = message.content[4:].split("/")
        await message.channel.send(vote[0])
        for i in range(1, len(vote)):
            choose = await message.channel.send(vote[i])
            await message.add_reaction(choose, '👍')

    if message.content.startswith("!레벨"):
        file = openpyxl.load_workbook("레벨.xlsx")
        sheet = file.active
        exp = [10, 20, 30, 40, 50]
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                sheet["B" + str(i)].value = sheet["B" + str(i)].value + 5
                if sheet["B" + str(i)].value >= exp[sheet["C" + str(i)].value - 1]:
                    sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1
                    await message.channel.send("레벨이 오름.\n현재 레벨 : " + str(sheet["C" + str(i)].value) + "\n경험치 : " + str(sheet["B" + str(i)].value))
                file.save("레벨.xlsx")
                break

            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = 0
                sheet["C" + str(i)].value = 1
                file.save("레벨.xlsx")
                break

            i += 1

    if message.content.startswith("!경고"):
        author = message.guild.get_member(int(message.content[4:22]))
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                file.save("경고.xlsx")
                if sheet["B" + str(i)].value == 3:
                    await message.guild.ban(author)
                    await message.channel.send("경고 3회 누적입니다. 서버에서 추방됩니다.")
                
                if sheet["B" + str(i)].value == 2:
                    await message.channel.send("경고 2회 누적입니다.")

                else:
                    await message.channel.send("경고를 1회 받았습니다.")
                break

            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(author.id)
                sheet["B" + str(i)].value = 1
                file.save("경고.xlsx")
                await message.channel.send("경고를 1회 받았습니다.")
                break
            i += 1

    if message.content.startswith("!해제"):
        author = message.guild.get_member(int(message.content[4:22]))
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(author.id):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) - 1
                file.save("경고.xlsx")
                await message.channel.send("경고 1회 해제했습니다.")
            break

access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
